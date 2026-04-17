from datetime import datetime
import csv
from io import BytesIO, StringIO
import json
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from sqlalchemy import asc, desc, func
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.models.mix import MixDesign, MixRevision, MixStatus
from app.schemas.mix import (
    IS_METHOD,
    MixDesignCreate,
    MixDesignOut,
    MixDesignUpdate,
    MixListResponse,
    RecalculateRequest,
    RecalculateResponse,
    RevisionOut,
)
from app.services.calculation import mix_snapshot, recalculate_mix
from app.services.exporter import export_csv, export_pdf, export_xlsx
from app.services.qr import generate_qr_assets


router = APIRouter(prefix="/mixes", tags=["mixes"])


def _apply_sort(query, sort_by: str, order: str):
    sort_map = {
        "recent": MixDesign.updated_at,
        "alphabetical": MixDesign.mix_name,
        "grade": MixDesign.concrete_grade,
        "mix_id": MixDesign.mix_id,
    }
    col = sort_map.get(sort_by, MixDesign.updated_at)
    return query.order_by(desc(col) if order == "desc" else asc(col))


def _ensure_qr_assets(
    mix: MixDesign,
    db: Session,
    force: bool = False,
    base_public_url: str | None = None,
) -> None:
    png_exists = bool(mix.qr_path_png) and Path(mix.qr_path_png).exists()
    svg_exists = bool(mix.qr_path_svg) and Path(mix.qr_path_svg).exists()
    if not force and png_exists and svg_exists:
        return

    png, svg = generate_qr_assets(mix.slug, base_public_url=base_public_url)
    mix.qr_path_png = png
    mix.qr_path_svg = svg
    db.commit()
    db.refresh(mix)


@router.get("", response_model=MixListResponse)
def list_mixes(
    db: Session = Depends(get_db),
    q: Optional[str] = None,
    grade: Optional[str] = None,
    design_method: Optional[str] = None,
    cement_type: Optional[str] = None,
    exposure_condition: Optional[str] = None,
    aggregate_size: Optional[float] = None,
    slump: Optional[float] = None,
    admixture_type: Optional[str] = None,
    project_tag: Optional[str] = None,
    category: Optional[str] = None,
    sort_by: str = "recent",
    order: str = "desc",
    page: int = 1,
    page_size: int = 20,
):
    query = db.query(MixDesign)

    if q:
        pattern = f"%{q}%"
        query = query.filter(
            MixDesign.mix_name.ilike(pattern)
            | MixDesign.mix_id.ilike(pattern)
            | MixDesign.remarks.ilike(pattern)
        )
    if grade:
        query = query.filter(MixDesign.concrete_grade == grade)
    if design_method:
        if design_method != IS_METHOD:
            raise HTTPException(status_code=400, detail=f"Only {IS_METHOD} is supported")
        query = query.filter(MixDesign.design_method == design_method)
    if cement_type:
        query = query.filter(MixDesign.cement_type == cement_type)
    if exposure_condition:
        query = query.filter(MixDesign.exposure_condition == exposure_condition)
    if aggregate_size is not None:
        query = query.filter(MixDesign.max_aggregate_size_mm == aggregate_size)
    if slump is not None:
        query = query.filter(MixDesign.slump_mm == slump)
    if admixture_type:
        query = query.filter(MixDesign.admixture_type == admixture_type)
    if project_tag:
        query = query.filter(MixDesign.project_tag == project_tag)
    if category:
        query = query.filter(MixDesign.category == category)

    total = query.count()
    query = _apply_sort(query, sort_by=sort_by, order=order)
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return MixListResponse(total=total, items=items)


@router.get("/dashboard/summary")
def dashboard_summary(db: Session = Depends(get_db)):
    total = db.query(func.count(MixDesign.id)).scalar() or 0
    approved = db.query(func.count(MixDesign.id)).filter(MixDesign.status == MixStatus.approved).scalar() or 0
    trial = db.query(func.count(MixDesign.id)).filter(MixDesign.status == MixStatus.trial).scalar() or 0
    archived = db.query(func.count(MixDesign.id)).filter(MixDesign.status == MixStatus.archived).scalar() or 0
    recent = db.query(MixDesign).order_by(desc(MixDesign.updated_at)).limit(8).all()
    return {
        "total_mixes": total,
        "approved": approved,
        "trial": trial,
        "archived": archived,
        "recent": [
            {"mix_id": m.mix_id, "mix_name": m.mix_name, "slug": m.slug, "grade": m.concrete_grade, "updated_at": m.updated_at}
            for m in recent
        ],
    }


@router.post("", response_model=MixDesignOut)
def create_mix(payload: MixDesignCreate, db: Session = Depends(get_db)):
    existing = db.query(MixDesign).filter((MixDesign.mix_id == payload.mix_id) | (MixDesign.slug == payload.slug)).first()
    if existing:
        raise HTTPException(status_code=400, detail="Mix ID or slug already exists")

    mix = MixDesign(**payload.model_dump())
    png, svg = generate_qr_assets(mix.slug)
    mix.qr_path_png = png
    mix.qr_path_svg = svg
    mix.download_ref = f"/api/mixes/{mix.slug}/export/pdf"
    db.add(mix)
    db.commit()
    db.refresh(mix)
    return mix


@router.get("/{slug}", response_model=MixDesignOut)
def get_mix(slug: str, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")
    return mix


@router.put("/{slug}", response_model=MixDesignOut)
def update_mix(slug: str, payload: MixDesignUpdate, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")

    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(mix, k, v)

    db.commit()
    db.refresh(mix)
    return mix


@router.delete("/{slug}")
def delete_mix(slug: str, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")
    db.delete(mix)
    db.commit()
    return {"message": "deleted"}


@router.post("/{slug}/duplicate", response_model=MixDesignOut)
def duplicate_mix(slug: str, db: Session = Depends(get_db)):
    src = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not src:
        raise HTTPException(status_code=404, detail="Mix not found")

    ts = datetime.utcnow().strftime("%H%M%S")
    new_mix_id = f"{src.mix_id}-COPY-{ts}"
    new_slug = f"{src.slug}-copy-{ts}"

    copy_data = {c.name: getattr(src, c.name) for c in src.__table__.columns}
    copy_data.pop("id", None)
    copy_data["mix_id"] = new_mix_id
    copy_data["slug"] = new_slug
    copy_data["mix_name"] = f"{src.mix_name} (Copy)"

    copy_mix = MixDesign(**copy_data)
    png, svg = generate_qr_assets(new_slug)
    copy_mix.qr_path_png = png
    copy_mix.qr_path_svg = svg
    copy_mix.download_ref = f"/api/mixes/{new_slug}/export/pdf"

    db.add(copy_mix)
    db.commit()
    db.refresh(copy_mix)
    return copy_mix


@router.post("/{slug}/recalculate", response_model=RecalculateResponse)
def recalculate(slug: str, payload: RecalculateRequest, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")

    old_value = getattr(mix, payload.parameter, None)
    outcome = recalculate_mix(mix, payload.parameter, payload.new_value)

    if payload.save_revision:
        rev = MixRevision(
            mix_design_id=mix.id,
            revision_label=f"rev-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            changed_parameter=payload.parameter,
            old_value=str(old_value),
            new_value=str(payload.new_value),
            warning_message="; ".join(outcome.warnings),
            snapshot_json=mix_snapshot(mix),
        )
        db.add(rev)

    db.commit()
    db.refresh(mix)

    return RecalculateResponse(updated_mix=mix, warnings=outcome.warnings)


@router.get("/{slug}/revisions", response_model=list[RevisionOut])
def get_revisions(slug: str, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")
    return db.query(MixRevision).filter(MixRevision.mix_design_id == mix.id).order_by(desc(MixRevision.created_at)).all()


@router.get("/{slug}/qr/png")
def get_qr_png(
    slug: str,
    db: Session = Depends(get_db),
    base_url: str | None = Query(default=None),
):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")
    _ensure_qr_assets(mix, db, force=True, base_public_url=base_url)
    return FileResponse(mix.qr_path_png, media_type="image/png", filename=f"{slug}.png")


@router.get("/{slug}/qr/svg")
def get_qr_svg(
    slug: str,
    db: Session = Depends(get_db),
    base_url: str | None = Query(default=None),
):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")
    _ensure_qr_assets(mix, db, force=True, base_public_url=base_url)
    return FileResponse(mix.qr_path_svg, media_type="image/svg+xml", filename=f"{slug}.svg")


@router.post("/qr/regenerate-all")
def regenerate_all_qr(db: Session = Depends(get_db)):
    mixes = db.query(MixDesign).all()
    for mix in mixes:
        png, svg = generate_qr_assets(mix.slug)
        mix.qr_path_png = png
        mix.qr_path_svg = svg
    db.commit()
    return {"updated": len(mixes)}


@router.get("/qr/sheet")
def qr_sheet(db: Session = Depends(get_db), limit: int = Query(20, ge=1, le=200)):
    mixes = db.query(MixDesign).order_by(asc(MixDesign.mix_id)).limit(limit).all()
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    y = 780
    x = 40
    col = 0

    for mix in mixes:
        if Path(mix.qr_path_png).exists():
            c.drawImage(mix.qr_path_png, x, y - 80, width=80, height=80)
            c.setFont("Helvetica", 8)
            c.drawString(x, y - 88, f"{mix.mix_id} / {mix.slug}")
            col += 1
            x += 170
            if col == 3:
                col = 0
                x = 40
                y -= 120
            if y < 140:
                c.showPage()
                y = 780
                x = 40
                col = 0

    c.save()
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=qr-sheet.pdf"})


@router.get("/{slug}/export/{fmt}")
def export_mix(slug: str, fmt: str, db: Session = Depends(get_db)):
    mix = db.query(MixDesign).filter(MixDesign.slug == slug).first()
    if not mix:
        raise HTTPException(status_code=404, detail="Mix not found")

    if fmt == "csv":
        payload = export_csv(mix)
        return Response(content=payload, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={slug}.csv"})
    if fmt == "xlsx":
        payload = export_xlsx(mix)
        return Response(content=payload, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename={slug}.xlsx"})
    if fmt == "pdf":
        payload = export_pdf(mix)
        return Response(content=payload, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={slug}.pdf"})

    raise HTTPException(status_code=400, detail="Unsupported format")


@router.post("/import")
async def import_mixes(file: UploadFile = File(...), db: Session = Depends(get_db)):
    name = file.filename.lower()
    rows: list[dict] = []

    if name.endswith(".csv"):
        data = (await file.read()).decode("utf-8")
        reader = csv.DictReader(StringIO(data))
        rows = [r for r in reader]
    elif name.endswith(".xlsx"):
        content = await file.read()
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({str(header[i]): row[i] for i in range(len(header))})
    else:
        raise HTTPException(status_code=400, detail="Only CSV/XLSX accepted")

    imported = 0
    for r in rows:
        if not r.get("mix_id") or not r.get("slug"):
            continue
        if db.query(MixDesign).filter((MixDesign.mix_id == r["mix_id"]) | (MixDesign.slug == r["slug"])).first():
            continue

        mix = MixDesign(
            mix_id=str(r["mix_id"]),
            slug=str(r["slug"]),
            mix_name=str(r.get("mix_name", "Imported Mix")),
            project_tag=str(r.get("project_tag", "imported")),
            concrete_grade=str(r.get("concrete_grade", "M30")),
            target_mean_strength=float(r.get("target_mean_strength", 38.0)),
            design_method=str(r.get("design_method", "IS 10262:2019")),
            cement_type=str(r.get("cement_type", "OPC 53")),
            max_aggregate_size_mm=float(r.get("max_aggregate_size_mm", 20.0)),
            exposure_condition=str(r.get("exposure_condition", "Moderate")),
            slump_mm=float(r.get("slump_mm", 100.0)),
            water_cement_ratio=float(r.get("water_cement_ratio", 0.45)),
            water_content_kg_m3=float(r.get("water_content_kg_m3", 180.0)),
            cement_content_kg_m3=float(r.get("cement_content_kg_m3", 400.0)),
            fine_agg_content_kg_m3=float(r.get("fine_agg_content_kg_m3", 680.0)),
            coarse_agg_content_kg_m3=float(r.get("coarse_agg_content_kg_m3", 1150.0)),
            admixture_type=str(r.get("admixture_type", "PCE Superplasticizer")),
            admixture_dosage_pct=float(r.get("admixture_dosage_pct", 0.8)),
            field_water_adjustment_kg=float(r.get("field_water_adjustment_kg", 0.0)),
            final_batch_water_kg=float(r.get("final_batch_water_kg", r.get("water_content_kg_m3", 180.0))),
            final_batch_cement_kg=float(r.get("final_batch_cement_kg", r.get("cement_content_kg_m3", 400.0))),
            final_batch_fine_agg_kg=float(r.get("final_batch_fine_agg_kg", r.get("fine_agg_content_kg_m3", 680.0))),
            final_batch_coarse_agg_kg=float(r.get("final_batch_coarse_agg_kg", r.get("coarse_agg_content_kg_m3", 1150.0))),
            mix_proportion_by_weight=str(r.get("mix_proportion_by_weight", "1:1.7:2.8 (w/c=0.45)")),
            assumptions=str(r.get("assumptions", "Imported assumptions")),
            remarks=str(r.get("remarks", "Imported")),
            category=str(r.get("category", "design mix")),
            download_ref=f"/api/mixes/{r['slug']}/export/pdf",
        )
        if mix.design_method != IS_METHOD:
            continue
        png, svg = generate_qr_assets(mix.slug)
        mix.qr_path_png = png
        mix.qr_path_svg = svg
        db.add(mix)
        imported += 1

    db.commit()
    return {"imported": imported, "total_rows": len(rows)}


@router.get("/database/export")
def export_database(db: Session = Depends(get_db)):
    mixes = db.query(MixDesign).order_by(asc(MixDesign.mix_id)).all()
    records = []
    for m in mixes:
        rec = {c.name: getattr(m, c.name) for c in m.__table__.columns}
        rec["status"] = str(m.status)
        rec["created_at"] = m.created_at.isoformat()
        rec["updated_at"] = m.updated_at.isoformat()
        records.append(rec)

    body = json.dumps(records, indent=2)
    return Response(content=body, media_type="application/json", headers={"Content-Disposition": "attachment; filename=mix-database.json"})
